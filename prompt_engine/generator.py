import os
import json
import requests
import boto3
import openai
import yaml
import fal_client
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
fal_key = os.environ.get('FAL_KEY')
openai.api_key = os.environ.get('OPENAI_API_KEY')
finetune_id = os.environ.get('FINE_TUNE_ID')
BUCKET_NAME = "falaipostings"

base_dir = os.path.dirname(os.path.abspath(__file__))
image_dir = os.path.join(base_dir, "images", "approved")
metadata_path = os.path.join(image_dir, "metadata.json")
excel_path = os.path.join(base_dir, "generated_images.xlsx")
prompt_file_path = os.path.join(base_dir, "prompt.yaml")
os.makedirs(image_dir, exist_ok=True)



if os.path.exists(metadata_path):
    os.remove(metadata_path)


with open(prompt_file_path, "r") as f:
    prompts = yaml.safe_load(f)
text_prompt = prompts["generate_post"]["user"]
system_prompt = prompts["generate_post"]["system"]
image_prompt = prompts["generate_image"]["prompt"]

response = openai.ChatCompletion.create(
    model="gpt-4",
    messages=[
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": text_prompt},
    ]
)
generated_text = response.choices[0].message.content.strip()
print("Text generated")

def on_queue_update(update):
    if isinstance(update, fal_client.InProgress):
        for log in update.logs:
            print(log["message"])

result = fal_client.subscribe(
    "fal-ai/flux-pro/v1.1-ultra-finetuned",
    arguments={
        "prompt": image_prompt,
        "guidance_scale": 10,
        "seed": 0,
        "sync_mode": False,
        "num_images": 3,
        "enable_safety_checker": False,
        "safety_tolerance": 6,
        "output_format": "jpeg",
        "aspect_ratio": "1:1",
        "finetune_id": finetune_id,
        "finetune_strength": 0.7
    },
    with_logs=True,
    on_queue_update=on_queue_update,
)
print("Images generated")
result["timestamp"] = datetime.now().isoformat()
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    if "Image" in wb.sheetnames:
        wb.remove(wb["Image"])
    ws1 = wb.create_sheet("Image")
    ws2 = wb["Data"] if "Data" in wb.sheetnames else wb.create_sheet("Data")
else:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Image"
    ws2 = wb.create_sheet("Data")
    ws2.append(["URL", "File Name", "Timings", "Prompt", "Generated Text"])

def style_button(cell, text, color, url):
    cell.value = text
    cell.hyperlink = url
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

image_row = 1
batch_metadata = {}

if result and 'images' in result:
    images = [img for img in result['images'] if img['url'].endswith('.jpeg') or img['url'].endswith('.jpg')]

    for idx, image_info in enumerate(images):
        image_url = image_info['url']
        file_id = image_url.split('/')[-1].split('.')[0]
        file_name = f"Aria_{result['seed']}_{file_id}.jpeg"
        file_path = os.path.join(image_dir, file_name)

        # Download image
        response = requests.get(image_url)
        if response.status_code != 200:
            print(f"Failed to download image: {response.status_code}")
            continue
        with open(file_path, 'wb') as f:
            f.write(response.content)

        # Add image to Excel
        img = XLImage(file_path)
        img.width = 793
        img.height = 497
        ws1.add_image(img, f"A{image_row}")

        # Display generated text below image (col N)
        ws1.cell(row=image_row, column=14).value = generated_text
        ws1.cell(row=image_row, column=14).alignment = Alignment(wrap_text=True)
        ws1.column_dimensions['N'].width = 65
        ws1.row_dimensions[image_row].height = 300

        approve_url = (
            f"https://dtvjy8dwkj.execute-api.ap-south-1.amazonaws.com/default/image_authenticator"
            f"?file={file_name}&url={image_url}&action=approve"
        )
        approve_cell = ws1.cell(row=image_row + 1, column=14)
        style_button(approve_cell, "APPROVE", "00B050", approve_url)

        row = ws2.max_row + 1
        ws2.cell(row=row, column=1).value = image_url
        ws2.cell(row=row, column=2).value = file_name
        ws2.cell(row=row, column=3, value=str(result.get("timestamp")))
        ws2.cell(row=row, column=4).value = result["prompt"]
        ws2.cell(row=row, column=5).value = generated_text

        batch_metadata[file_name] = {
            "url": image_url,
            "text": generated_text
        }

        image_row += 27

with open(metadata_path, "w") as f:
    json.dump(batch_metadata, f)

s3 = boto3.client("s3")
s3.upload_file(metadata_path, BUCKET_NAME, "temp/metadata.json")
print("Metadata sent")
wb.save(excel_path)
print(f"Excel updated")
os.utime("prompt.yaml", None)